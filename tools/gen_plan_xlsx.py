# -*- coding: utf-8 -*-
"""生成《个人网站改良升级工作计划》xlsx（一次性规划脚本，可重跑覆盖）"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\alext\WorkBuddy\2026-07-23-16-40-27\个人网站改良升级工作计划-v1.0.0.xlsx"

HEAD_FILL = PatternFill("solid", fgColor="2B1B4A")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=11, name="Microsoft YaHei")
CELL_FONT = Font(size=10, name="Microsoft YaHei")
TITLE_FONT = Font(size=13, bold=True, color="2B1B4A", name="Microsoft YaHei")
P0_FILL = PatternFill("solid", fgColor="FFD6E0")
P1_FILL = PatternFill("solid", fgColor="FFF3CD")
P2_FILL = PatternFill("solid", fgColor="E2F0D9")
THIN = Side(style="thin", color="BFB4D6")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_sheet(ws, headers, rows, widths, title, prio_col=None, wrap_cols=()):
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.row_dimensions[1].height = 26

    hr = 3
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=hr, column=c, value=h)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[hr].height = 22

    for r, row in enumerate(rows, hr + 1):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = CELL_FONT
            cell.border = BORDER
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=(c in wrap_cols),
                horizontal="left",
            )
        if prio_col is not None:
            p = str(row[prio_col - 1])
            fill = {"P0": P0_FILL, "P1": P1_FILL, "P2": P2_FILL}.get(p)
            if fill:
                ws.cell(row=r, column=prio_col).fill = fill
                ws.cell(row=r, column=prio_col).alignment = Alignment(
                    horizontal="center", vertical="top"
                )

    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.freeze_panes = ws.cell(row=hr + 1, column=1)
    ws.auto_filter.ref = "A%d:%s%d" % (
        hr,
        get_column_letter(len(headers)),
        hr + len(rows),
    )


wb = Workbook()

# ---------- Sheet1 审计缺陷清单 ----------
ws1 = wb.active
ws1.title = "审计缺陷清单"
h1 = ["编号", "优先级", "缺陷", "现状证据", "业务影响", "修复方案", "对应任务"]
rows1 = [
    ["D1", "P0", "询盘表单无后端，提交不发送任何数据",
     "contact.html 表单 submit 仅弹出「静态站演示」提示；main.js 第 12 行写死提示文案",
     "B2B 招商模块做得完整，但一条线索都进不来，漏斗只有上口没有下口",
     "表单工程化：前端校验 + honeypot 反垃圾 + 多通道提交（FormSubmit 邮件 / 自定义 webhook）", "T2"],
    ["D2", "P0", "缺品牌矩阵 / 战绩 / 招聘三类承载页",
     "站点仅 index、about、work、blog、contact 5 个主页面",
     "B 端访客看不到集团 9 品牌实力与可核验战绩，信任建立成本高；招聘无入口",
     "新建 brands.html / achievements.html / careers.html，接全站导航", "T4-T6"],
    ["D3", "P0", "海外访客默认中文，需手动切换",
     "i18n.js: current = localStorage.getItem || 'zh'，不做浏览器语言判断",
     "B2B 场景主力是海外客户，打开即中文会直接损失询盘",
     "首次访问按 navigator.language 自动判定，手动切换优先", "T9"],
    ["D4", "P0", "全站无 schema.org 结构化数据",
     "仅 index 页有基础 og:* meta，无一处 JSON-LD",
     "搜索结果没有富摘要（站点名/面包屑/作者），点击率受损",
     "新增 js/seo.js 按页面类型注入 Organization/Person/Article/BreadcrumbList", "T8"],
    ["D5", "P1", "无访问统计",
     "全站未引入任何统计脚本",
     "无法判断哪些文章引流、B2B 页有没有人看，功能决策靠拍脑袋",
     "配置化统计接入位（GoatCounter/Clarity），未配置不加载任何脚本", "T11"],
    ["D6", "P1", "无内容沉淀渠道（RSS / 订阅）",
     "15 篇文章无 RSS 源",
     "内容无复访、无私域沉淀，清仓/招商内容缺直达渠道",
     "tools/gen_rss.py 生成 rss.xml，页脚与博客页加订阅入口", "T10"],
    ["D7", "P1", "无评论互动",
     "文章页无评论区",
     "缺停留时长与互动信号，站点显得是「死站」",
     "Giscus 配置化挂载（GitHub Discussions 驱动，免费无广告）", "T11"],
    ["D8", "P1", "表单无校验、无成功态、无障碍属性不全",
     "仅 required 属性，label 绑定不完整，无提交中/失败态",
     "垃圾提交风险，移动端体验差，无障碍不达标",
     "邮箱格式校验 + 防重复提交 + 双语成功/失败态 + aria 属性", "T2"],
    ["D9", "P2", "无 PWA（manifest / Service Worker）",
     "无 manifest.webmanifest、无 sw.js、无应用图标",
     "移动端无法「装」到桌面，弱网/离线直接白屏",
     "新增 manifest + sw（HTML 网络优先、静态资源缓存优先）+ 192/512 图标", "T12"],
    ["D10", "P2", "无品牌默认分享图",
     "index 页 og:image 借用 assets/img/work-5-agnes.jpg，其余页面无 og meta",
     "文章分享到 LinkedIn / WhatsApp 无图无摘要，专业感弱",
     "生成 1200×630 品牌 OG 图，全站补齐 og:* 与 twitter:card", "T13"],
    ["D11", "P2", "无电子名片与二维码",
     "contact 页仅邮箱与 GitHub 链接",
     "线下展会/拜访场景无法快速留存联系方式",
     "生成 vCard 下载 + 微信二维码占位", "T14"],
    ["D12", "P2", "访客疑问无人即时解答",
     "全站无客服/问答入口",
     "「怎么代理」「做什么品牌」这类高频疑问无人答即流失",
     "规则型 AI 问答悬浮窗（中英双语关键词匹配，未命中引导留资）", "T15"],
    ["D13", "P1", "缓存版本参数不统一",
     "index 为 v1.5.5，contact 页 style.css?v=1.4.5、i18n.js?v=1.5.1",
     "部分页面可能长期加载旧版 CSS/JS，导致「改了没生效」",
     "全站统一升级到 v1.6.0，后续每次改动必须同步 bump", "T7"],
]
style_sheet(ws1, h1, rows1, [6, 8, 30, 40, 34, 40, 10],
            "一、现状审计：缺陷清单（13 项）", prio_col=2, wrap_cols=(3, 4, 5, 6))

# ---------- Sheet2 工作计划 ----------
ws2 = wb.create_sheet("工作计划")
h2 = ["阶段", "任务号", "任务", "交付物", "可否全自动化", "依赖项", "验收标准"]
rows2 = [
    ["S1 商业闭环", "T1", "输出改良升级工作计划方案", "个人网站改良升级工作计划-v1.0.0.xlsx", "是", "无", "4 张表、16 项任务、依赖项可追溯"],
    ["S1 商业闭环", "T2", "询盘表单工程化改造", "js/site-config.js、js/inquiry.js、contact.html 改造", "是",
     "FormSubmit 邮箱（可选，不填则走 webhook）", "校验生效、双语提示、防重复提交、提交不 404"],
    ["S1 商业闭环", "T3", "交付钉钉表单转发器", "tools/dingtalk-form-relay.js + 部署指引", "是（部署需用户10分钟）",
     "Cloudflare 账号 + 钉钉机器人 webhook", "Worker 代码可直接粘贴部署，CORS 与鉴权齐备"],
    ["S1 商业闭环", "T4", "新建品牌矩阵页", "brands.html", "是", "无", "9 品牌卡片、双语切换正常、移动端不溢出"],
    ["S1 商业闭环", "T5", "新建成绩单页", "achievements.html", "是", "战绩截图（可选）", "数据卡+时间线、双语、占位标注清晰"],
    ["S1 商业闭环", "T6", "新建招聘页", "careers.html", "是", "岗位 JD 终稿（可选）", "岗位卡片、投递入口可达、双语"],
    ["S2 流量基础", "T7", "导航/页脚/sitemap 接入 + 版本统一", "全站 HTML、sitemap.xml", "是", "无",
     "新页全部进导航、sitemap 收录、所有页面缓存参数 = v1.6.0"],
    ["S2 流量基础", "T8", "注入 JSON-LD 结构化数据", "js/seo.js + 全站引入", "是", "无",
     "首页 Organization/Person，文章页 Article，Google 富媒体测试可解析"],
    ["S2 流量基础", "T9", "访客语言自动检测", "js/i18n.js 改造", "是", "无", "首次访问英文浏览器显示英文，手动切换后记忆"],
    ["S2 流量基础", "T10", "生成 RSS 订阅源", "rss.xml、tools/gen_rss.py、页脚入口", "是", "无", "rss.xml 含 15 篇、XML 可被校验器解析"],
    ["S3 体验品牌", "T11", "统计与评论配置化接入位", "js/analytics.js、js/comments.js", "是",
     "GoatCounter/Clarity ID、Giscus repo（可选）", "未配置时零脚本加载、配置后一行生效"],
    ["S3 体验品牌", "T12", "PWA 离线支持", "manifest.webmanifest、sw.js、图标 192/512", "是", "无", "Chrome 可安装、离线回退页可用、SW 版本可更新"],
    ["S3 体验品牌", "T13", "OG 默认图 + 社交 meta 补齐", "assets/img/og-default.png、全站 meta", "是", "无", "分享卡片带图带摘要、og:image 无 404"],
    ["S3 体验品牌", "T14", "电子名片 vCard + 二维码位", "assets/戴程鹏.vcf、contact.html 增强", "是", "微信二维码图片（可选）", "vCard 可导入通讯录、二维码位有明显待补提示"],
    ["S3 体验品牌", "T15", "AI 问答助手", "js/assistant.js + CSS", "是", "无", "中英关键词命中、未命中引导留资、移动端不遮挡内容"],
    ["S4 发布验收", "T16", "冒烟测试 + 提交发布 + 验收报告", "验收报告、git commit & push", "是", "GitHub 凭据",
     "无 404、无控制台报错、GitHub Pages 线上可访问"],
]
style_sheet(ws2, h2, rows2, [13, 8, 28, 34, 18, 26, 38],
            "二、改良升级工作计划：16 项任务（分 4 阶段顺序执行）", prio_col=None, wrap_cols=(3, 4, 6, 7))

# ---------- Sheet3 需你提供 ----------
ws3 = wb.create_sheet("需你提供的配置")
h3 = ["配置项", "用途", "怎么拿", "不提供的后果", "填写位置", "状态"]
rows3 = [
    ["接收询盘的邮箱（推荐先填）", "访客提交表单后直接进你邮箱", "任意邮箱即可，FormSubmit 首次提交需点确认邮件激活",
     "表单只能走自定义 webhook 或本地提示，线索仍收不到", "js/site-config.js → FORM_ENDPOINT", "待填"],
    ["钉钉机器人 webhook（可选）", "询盘实时推送到钉钉群", "钉钉群设置 → 群机器人 → 自定义 → 复制 Webhook",
     "有询盘不能秒级提醒，需手动查邮箱", "Cloudflare Worker 环境变量 DINGTALK_WEBHOOK", "待填"],
    ["Cloudflare 账号（可选）", "托管表单转发器（免费额度够用）", "cloudflare.com 免费注册，Workers 免费层 10 万请求/日",
     "只能用邮件通道接收询盘", "—", "待填"],
    ["战绩/产品截图（可选）", "成绩单页与品牌页配图", "亚马逊品牌榜截图、产品图",
     "页面用占位框，视觉说服力打折", "assets/img/ 目录", "待补"],
    ["微信二维码（可选）", "访客扫码加微信", "微信 → 我 → 二维码 → 保存图片",
     "二维码位显示「待补充」占位", "assets/img/wechat-qr.png", "待补"],
    ["统计站点 ID（可选）", "看访问量与热门文章", "GoatCounter 免费注册拿 sitecode，或 Clarity 拿 project id",
     "网站无访问数据，只能盲改", "js/site-config.js → ANALYTICS", "待填"],
    ["Giscus 配置（可选）", "文章页评论区", "GitHub 仓库开 Discussions + giscus.app 生成 repoId/categoryId",
     "文章无评论互动", "js/site-config.js → GISCUS", "待填"],
    ["Google Search Console 验证（建议做）", "让搜索引擎收录并看关键词排名", "search.google.com 添加域名，下载验证文件上传到仓库根",
     "站点长期不被索引，博客等于自嗨", "仓库根目录验证 html", "待操作"],
    ["岗位 JD 终稿（可选）", "招聘页内容", "HR 或你提供", "页面用通用模板 + 待确认标注", "careers.html", "待确认"],
]
style_sheet(ws3, h3, rows3, [26, 26, 44, 34, 34, 8],
            "三、需要你提供 / 确认的配置项（不阻塞自动执行，填一行即生效）", prio_col=None, wrap_cols=(1, 2, 3, 4, 5))

# ---------- Sheet4 验收清单 ----------
ws4 = wb.create_sheet("验收清单")
h4 = ["类别", "验收项", "验证方法", "预期结果", "结果"]
rows4 = [
    ["页面", "3 个新页面可访问", "本地 HTTP 逐个打开 brands/achievements/careers", "200，无 404，无控制台报错", "待验"],
    ["页面", "全站导航链接无死链", "遍历所有页面所有 a[href] 检查目标存在", "0 个死链", "待验"],
    ["双语", "中英切换覆盖新增内容", "点 EN 按钮遍历新页与表单", "新页面文本与 placeholder 全部切换", "待验"],
    ["双语", "英文浏览器首访即英文", "Puppeteer 设置 navigator.language=en-US 首访", "首屏为英文，切换后记忆", "待验"],
    ["表单", "前端校验拦截非法输入", "提交空表单 / 错误邮箱", "对应提示，不发起请求", "待验"],
    ["表单", "提交流程不 404", "配置 FormSubmit 邮箱后真实提交一次", "收到邮件，页面显示成功态", "待验"],
    ["SEO", "JSON-LD 可解析", "查看页面 source 中的 application/ld+json", "首页 Organization+Person，文章页 Article", "待验"],
    ["SEO", "sitemap 收录新页", "打开 sitemap.xml", "含 3 个新页与 15 篇文章", "待验"],
    ["SEO", "RSS 可解析", "XML 校验器打开 rss.xml", "含 15 条 item，无解析错误", "待验"],
    ["PWA", "可安装且离线可用", "Chrome DevTools Application 面板", "manifest 有效、SW 激活、离线有回退", "待验"],
    ["视觉", "移动端不溢出", "375px 视口截图", "无横向滚动，卡片堆叠正常", "待验"],
    ["视觉", "新页面与首页风格一致", "截图比对配色与卡片样式", "沿用紫黑霓虹主题", "待验"],
    ["性能", "首屏加载正常", "本地计时 + Network 面板", "hero 视频 1.5MB 内，无阻塞资源", "待验"],
    ["发布", "GitHub Pages 线上生效", "推送后访问线上地址", "线上与本地一致", "待验"],
]
style_sheet(ws4, h4, rows4, [10, 28, 40, 40, 8],
            "四、验收清单（14 项，冒烟测试逐条勾选）", prio_col=None, wrap_cols=(2, 3, 4))

os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print("SAVED:", OUT)
